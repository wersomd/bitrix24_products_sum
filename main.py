import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = "products.xlsx"

df = pd.read_excel(file_path, engine="openpyxl")

purchase_price_col = "Закупочная цена"
available_quantity_col = "Доступный остаток"


res = []


for i in range(len(df)):
    
    if df[available_quantity_col][i] == 0:
        purchase_price = str(df[purchase_price_col][i]).replace(" тенге", "").replace(" ", "").replace("\u00A0", "")
        q3 = float(purchase_price)
    
    else:
        purchase_price = str(df[purchase_price_col][i]).replace(" тенге", "").replace(" ", "").replace("\u00A0", "")
        q3 = float(purchase_price) * float(df[available_quantity_col][i])

    res.append(q3)

df["Итоговая стоимость"] = res
df["Итоговая стоимость (тенге)"] = df["Итоговая стоимость"].apply(
    lambda x: f"{float(str(x).replace(' тенге', '').replace('\xa0', '')):,.2f} тенге" 
    if pd.notna(x) and isinstance(x, (int, float, str)) 
    and str(x).replace(".", "", 1).isdigit() 
    else None
)

df.drop(columns=["Итоговая стоимость"], inplace=True)


output_path = "output.xlsx"
df.to_excel(output_path, index=False, engine="openpyxl")

wb = load_workbook(output_path)
ws = wb.active

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(output_path)
print(f"Результат сохранён в файл: {output_path}")
